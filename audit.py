"""
Katalog ↔ Excel uyum raporu.
PDF'i tarayarak şu sorunları tespit eder:
  - Tipografik hatalı kodlar (OCR-confusion: 0↔O, 1↔I vb.)
  - Başlık konumu olmayan kodlar (sadece tabloda → fiyat alamaz)
  - Bir kodun tablo satırında farklı, başlık konumunda farklı yazılması (tutarsızlık)
  - Aynı sayfadaki dik üst-üste yığın kodlar
  - Excel'deki ama PDF'te hiç görünmeyen kodlar
  - Çoklu sayfada dağıtılmış varyantlar
"""
from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

from extract_codes import _pick_typo_winner
from inject_prices import CODE_RE, _is_table_row, find_code_positions

HERE = Path(__file__).parent
PDF = HERE / "MD CATALOG 2025 TR-EN.pdf"
XLSX = HERE / "fiyat-listesi.xlsx"
OUT = HERE / "audit-rapor.txt"


def main():
    # Excel kodlarını yükle
    wb = load_workbook(XLSX)
    ws = wb.active
    excel_codes: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            excel_codes.add(str(row[0]).strip())

    # PDF'i tara
    code_positions: dict[str, list[tuple[int, bool]]] = defaultdict(list)  # code -> [(page, is_table)]
    raw_pdf_codes: set[str] = set()
    stack_warnings: list[str] = []
    page_titles: dict[int, list[str]] = defaultdict(list)

    with pdfplumber.open(PDF) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for m in CODE_RE.finditer(text):
                raw_pdf_codes.add(m.group(0))

            words = list(page.extract_words())
            positions = find_code_positions(page)
            seen_codes_on_page = set()
            for pos in positions:
                code = pos[0]
                is_tbl = _is_table_row(pos, words)
                code_positions[code].append((i, is_tbl))
                if not is_tbl:
                    if code not in seen_codes_on_page:
                        page_titles[i].append(code)
                        seen_codes_on_page.add(code)

            # Aynı sayfadaki yığın algılama (iç içe başlıklar)
            titles_on_page = [p for p in positions if not _is_table_row(p, words)]
            for i_a, a in enumerate(titles_on_page):
                _, ax0, ay_top, ax1, ay_bot = a
                for b in titles_on_page[i_a + 1:]:
                    _, bx0, by_top, bx1, by_bot = b
                    if min(ax1, bx1) - max(ax0, bx0) > 0 and -10 <= (by_top - ay_bot) <= 20:
                        stack_warnings.append(f"  sayfa {i}: '{a[0]}' ÜSTÜ + '{b[0]}' ALT (yığın)")

    # Kodları kategorize et
    codes_with_title: set[str] = set()
    codes_table_only: set[str] = set()
    for code, occ in code_positions.items():
        if any(not is_tbl for _, is_tbl in occ):
            codes_with_title.add(code)
        else:
            codes_table_only.add(code)

    # OCR typo'ları tespit (PDF içinde kanonik-typo çiftleri)
    pdf_codes_list = sorted(raw_pdf_codes)
    typo_pairs: list[tuple[str, str]] = []
    for i, a in enumerate(pdf_codes_list):
        for b in pdf_codes_list[i + 1:]:
            winner = _pick_typo_winner(a, b)
            if winner:
                canonical, typo = winner
                typo_pairs.append((canonical, typo))

    # Excel ↔ PDF tutarsızlık
    excel_minus_pdf = excel_codes - raw_pdf_codes
    pdf_minus_excel = raw_pdf_codes - excel_codes

    # ÇOK SAYFAYA YAYILMIŞ kodlar (1'den fazla sayfada başlık)
    multi_page_titles = {
        c: sorted({p for p, is_tbl in occ if not is_tbl})
        for c, occ in code_positions.items()
        if len({p for p, is_tbl in occ if not is_tbl}) > 1
    }

    # Raporu yaz
    lines: list[str] = []
    lines.append("=" * 70)
    lines.append("KATALOG ↔ EXCEL AUDIT RAPORU")
    lines.append("=" * 70)
    lines.append("")
    lines.append(f"PDF dosyası: {PDF.name}")
    lines.append(f"Excel dosyası: {XLSX.name}")
    lines.append(f"PDF'teki unique kod (raw): {len(raw_pdf_codes)}")
    lines.append(f"PDF'te başlık konumlu kod: {len(codes_with_title)}")
    lines.append(f"PDF'te SADECE tabloda olan kod: {len(codes_table_only)}")
    lines.append(f"Excel'deki kod: {len(excel_codes)}")
    lines.append("")

    # 1. PDF'teki OCR typo'lar
    lines.append("─" * 70)
    lines.append(f"1. PDF'TE OCR TYPO'LARI (kanonik ↔ typo) — {len(typo_pairs)} adet")
    lines.append("─" * 70)
    lines.append("Bu kodlar dedup mantığı ile elenir. PDF'te elle düzeltirsen daha net olur.")
    lines.append("")
    if typo_pairs:
        for canonical, typo in typo_pairs:
            pages_canon = sorted({p for p, _ in code_positions.get(canonical, [])})
            pages_typo = sorted({p for p, _ in code_positions.get(typo, [])})
            lines.append(f"  '{canonical}' (sayfa {pages_canon}) ↔ TYPO '{typo}' (sayfa {pages_typo})")
    else:
        lines.append("  Yok.")
    lines.append("")

    # 2. Başlık konumu olmayan kodlar (sadece tabloda)
    lines.append("─" * 70)
    lines.append(f"2. BAŞLIK KONUMU OLMAYAN KODLAR — {len(codes_table_only)} adet")
    lines.append("─" * 70)
    lines.append("Bu kodlar sadece spec tablosunda görünüyor; ayrı ürün başlığı yok.")
    lines.append("Bu kodlara fiyat BASILMAYACAK. Eğer fiyat istiyorsan PDF'e ürün başlığı eklenmeli.")
    lines.append("")
    if codes_table_only:
        for code in sorted(codes_table_only):
            pages = sorted({p for p, _ in code_positions[code]})
            lines.append(f"  '{code}' — sayfa {pages}")
    else:
        lines.append("  Yok.")
    lines.append("")

    # 3. Üst üste yığın kodlar
    lines.append("─" * 70)
    lines.append(f"3. ÜST ÜSTE YIĞIN KODLAR — {len(stack_warnings)} adet")
    lines.append("─" * 70)
    lines.append("İki kod yatay olarak iç içe; algoritma birinin fiyatını üste, ötekinin altına basar.")
    lines.append("")
    if stack_warnings:
        lines.extend(stack_warnings)
    else:
        lines.append("  Yok.")
    lines.append("")

    # 4. Birden fazla sayfada başlık olarak görünen kodlar
    lines.append("─" * 70)
    lines.append(f"4. ÇOK SAYFAYA YAYILMIŞ KODLAR — {len(multi_page_titles)} adet")
    lines.append("─" * 70)
    lines.append("Aynı kod, farklı sayfalarda ürün başlığı olarak görünüyor (varyant/duplikat).")
    lines.append("")
    if multi_page_titles:
        for code, pages in sorted(multi_page_titles.items()):
            lines.append(f"  '{code}' — sayfa {pages}")
    else:
        lines.append("  Yok.")
    lines.append("")

    # 5. Excel'de var, PDF'te yok
    lines.append("─" * 70)
    lines.append(f"5. EXCEL'DE VAR, PDF'TE HİÇ YOK — {len(excel_minus_pdf)} adet")
    lines.append("─" * 70)
    lines.append("Excel'e yazıldı ama PDF'te bu kod hiç geçmiyor. Yazım hatası kontrol et.")
    lines.append("")
    if excel_minus_pdf:
        for code in sorted(excel_minus_pdf):
            lines.append(f"  '{code}'")
    else:
        lines.append("  Yok.")
    lines.append("")

    # 6. PDF'te var, Excel'de yok
    lines.append("─" * 70)
    lines.append(f"6. PDF'TE VAR, EXCEL'DE YOK — {len(pdf_minus_excel)} adet")
    lines.append("─" * 70)
    lines.append("PDF'te var ama Excel şablonunda yok. (Çoğu OCR typo olur — bkz. bölüm 1.)")
    lines.append("")
    if pdf_minus_excel:
        for code in sorted(pdf_minus_excel):
            pages = sorted({p for p, _ in code_positions.get(code, [])})
            lines.append(f"  '{code}' — sayfa {pages}")
    else:
        lines.append("  Yok.")
    lines.append("")

    # 7. Sayfa bazlı özet
    lines.append("─" * 70)
    lines.append("7. SAYFA BAZLI BAŞLIK SAYIMI")
    lines.append("─" * 70)
    lines.append("Hangi sayfada kaç ürün başlığı var (fiyat basılacak adet).")
    lines.append("")
    for p in sorted(page_titles):
        codes = page_titles[p]
        lines.append(f"  sayfa {p:3}: {len(codes)} ürün → {', '.join(codes)}")
    lines.append("")

    lines.append("=" * 70)
    lines.append("RAPOR BİTTİ")
    lines.append("=" * 70)

    OUT.write_text("\n".join(lines), encoding="utf-8")
    # Konsola da kısa özet
    print(f"✓ Rapor yazıldı: {OUT.name} ({len(lines)} satır)")
    print()
    print(f"  OCR typo:                  {len(typo_pairs)}")
    print(f"  Sadece tabloda olan kod:   {len(codes_table_only)}")
    print(f"  Üst üste yığın:            {len(stack_warnings)}")
    print(f"  Çok sayfaya yayılmış:      {len(multi_page_titles)}")
    print(f"  Excel'de var, PDF'te yok:  {len(excel_minus_pdf)}")
    print(f"  PDF'te var, Excel'de yok:  {len(pdf_minus_excel)}")


if __name__ == "__main__":
    main()
