"""
M&D Katalog Fiyat Generator — Web Arayüzü
==========================================
Çok markalı katalog için fiyat yönetim ve PDF üretim uygulaması.

Markalar:
  • İNOKSAN     (ZMD-XXX)   → İnoksan listesi
  • EMPERO      (ZMD.XXX)   → Empero listesi (opsiyonel)
  • ATALAY      (MD.7AXX)   → M&D Prime listesi
  • SEAMAC      (MD.XXX)    → Seamac listesi (opsiyonel)
  • ÖZTİRYAKİLER (MDK4/MDE4) → Öztiryakiler listesi + iskonto/marj hesabı

Çalıştırma: streamlit run streamlit_app.py
"""
from __future__ import annotations

import io
import re
import tempfile
from collections import Counter, defaultdict
from pathlib import Path

import pdfplumber
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extract_codes import _dedupe_typos, _pick_typo_winner
from inject_prices import (
    CODE_RE,
    apply_fuzzy_matches,
    build_overlay,
    find_code_positions,
    _is_table_row,
    load_prices,
    merge,
)

# ───────────────────────────── Sayfa ayarı ─────────────────────────────
st.set_page_config(
    page_title="M&D Katalog Fiyat Generator",
    page_icon="📕",
    layout="wide",
)
st.title("📕 M&D Katalog Fiyat Generator")
st.caption("PDF katalog + marka fiyat listelerinden tek tıkla fiyatlı katalog üretir.")

# ───────────────────────────── Yardımcılar ─────────────────────────────
BRAND_COLORS = {
    "İNOKSAN":      "#FFE699",
    "EMPERO":       "#B4C7E7",
    "ATALAY":       "#C6E0B4",
    "SEAMAC":       "#F4B084",
    "ÖZTİRYAKİLER": "#9BC2E6",
}

PREFIX_PRIORITY = [
    "79K4", "79E4", "72K4", "72E4",
    "79K3", "79E3", "72K3", "72E3",
    "7919", "7219",
]


def brand_of(code: str) -> str:
    if code.startswith("MDK") or code.startswith("MDE"):
        return "ÖZTİRYAKİLER"
    if code.startswith("ZMD."):
        return "EMPERO"
    if code.startswith("ZMD-"):
        return "İNOKSAN"
    if code.startswith("MD.7A"):
        return "ATALAY"
    if code.startswith("MD."):
        return "SEAMAC"
    return "?"


def super_norm(code: str) -> str:
    s = code.upper().strip()
    for p in ["INO-ZMD-", "INO-MD-", "INO-", "ZMD-", "ZMD.", "MD.", "MD-"]:
        if s.startswith(p):
            s = s[len(p):]
            break
    return re.sub(r"[\-\.\s/]", "", s)


def _write_temp(data: bytes, suffix: str) -> Path:
    tf = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    tf.write(data)
    tf.close()
    return Path(tf.name)


# ───────────────────── Katalog tarama (cache'li) ─────────────────────
@st.cache_data(show_spinner=False)
def scan_catalog(pdf_bytes: bytes):
    """PDF'i tarayıp kodları, sayfaları, audit istatistiklerini döner."""
    pdf_path = _write_temp(pdf_bytes, ".pdf")
    codes_pages = {}
    code_positions = defaultdict(list)
    stacks = []
    n_pages = 0
    with pdfplumber.open(pdf_path) as pdf:
        n_pages = len(pdf.pages)
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for m in CODE_RE.finditer(text):
                code = m.group(0)
                if code not in codes_pages:
                    codes_pages[code] = i
            words = list(page.extract_words())
            positions = find_code_positions(page)
            titles_on_page = []
            seen = set()
            for pos in positions:
                c = pos[0]
                is_tbl = _is_table_row(pos, words)
                code_positions[c].append((i, is_tbl))
                if not is_tbl and c not in seen:
                    seen.add(c)
                    titles_on_page.append(pos)
            for ia, a in enumerate(titles_on_page):
                _, ax0, ay_top, ax1, ay_bot = a
                for b in titles_on_page[ia + 1:]:
                    _, bx0, by_top, bx1, by_bot = b
                    if min(ax1, bx1) - max(ax0, bx0) > 0 and -10 <= (by_top - ay_bot) <= 20:
                        stacks.append((i, a[0], b[0]))
    codes_pages = _dedupe_typos(codes_pages)
    sorted_codes = sorted(codes_pages.keys())
    typo_pairs = []
    for i, a in enumerate(sorted_codes):
        for b in sorted_codes[i + 1:]:
            w = _pick_typo_winner(a, b)
            if w:
                typo_pairs.append(w)
    table_only = sorted([c for c, occ in code_positions.items() if all(t for _, t in occ)])
    return {
        "codes_pages": codes_pages,
        "n_pages": n_pages,
        "typo_pairs": typo_pairs,
        "table_only": table_only,
        "stacks": stacks,
        "pdf_path": pdf_path,
    }


# ─────────────── Marka listesi okuma / eşleştirme ───────────────
def index_brand_list(xlsx_bytes: bytes, brand: str):
    """Marka fiyat listesinden kod → fiyat sözlüğü oluştur."""
    p = _write_temp(xlsx_bytes, ".xlsx")
    wb = load_workbook(p, data_only=True)
    by_norm: dict = {}
    oz_by_sig: dict = {}
    if brand == "ÖZTİRYAKİLER":
        # Öztiryakiler — multi-sheet, signature bazlı
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                code = str(row[0]).strip()
                if re.match(r"^[0-9A-Z]{4}\.[0-9A-Z]+\.[0-9A-Z]+$", code):
                    parts = code.split(".")
                    sig = (parts[1], parts[2])
                    name = str(row[1] or "")
                    price = row[2]
                    oz_by_sig.setdefault(sig, []).append((code, name, price))
        return {"oz_by_sig": oz_by_sig}
    else:
        # İnoksan / Atalay / Empero / Seamac — düz liste
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            code = str(row[1]).strip()
            name = str(row[2] or "")
            price = row[3]
            curr = row[5] if len(row) > 5 else "EUR"
            by_norm[super_norm(code)] = (code, name, price, curr or "EUR")
        return {"by_norm": by_norm}


def oz_match(md_code: str, oz_by_sig: dict):
    parts = md_code.split(".")
    if len(parts) != 3:
        return None
    _, middle, suffix = parts
    sigs = []
    if "MDB" in middle:
        sigs = [(middle.replace("MDB", "NMV"), suffix),
                (middle.replace("MDB", "NTS"), suffix)]
    elif "MDD" in middle:
        sigs = [(middle.replace("MDD", "LMV"), suffix),
                (middle.replace("MDD", "LTS"), suffix)]
    elif "PZC" in middle:
        sigs = [(middle, suffix)]
    else:
        return None
    cands = []
    for s in sigs:
        if s in oz_by_sig:
            cands.extend(oz_by_sig[s])
    if not cands:
        return None
    for p in PREFIX_PRIORITY:
        for c in cands:
            if c[0].startswith(p + "."):
                return c
    return cands[0]


# ────────────────────── Excel hazırlama ──────────────────────
def build_price_excel(
    codes_pages: dict,
    brand_data: dict,
    oz_discount: float,
    oz_otv: float,
    oz_margin: float,
) -> bytes:
    """fiyat-listesi.xlsx için bytes döner."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fiyat Listesi"
    hdrs = ["Marka", "Ürün Kodu", "Fiyat", "Para Birimi", "Açıklama",
            "Sayfa (PDF)", "Kaynak"]
    ws.append(hdrs)
    h_font = Font(bold=True, color="FFFFFF")
    h_fill = PatternFill("solid", fgColor="305496")
    for c in ws[1]:
        c.font = h_font
        c.fill = h_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    sorted_codes = sorted(codes_pages.items(), key=lambda x: (brand_of(x[0]), x[1], x[0]))
    for code, page in sorted_codes:
        brand = brand_of(code)
        price = None
        currency = "EUR"
        desc = ""
        source = ""

        if brand == "ÖZTİRYAKİLER" and "ÖZTİRYAKİLER" in brand_data:
            m = oz_match(code, brand_data["ÖZTİRYAKİLER"]["oz_by_sig"])
            if m:
                oz_code, name, oz_price = m
                price = round(oz_price * oz_discount * oz_otv * oz_margin, 2)
                currency = "EUR"
                desc = name
                source = f"ÖZTİRYAKİLER {oz_code}"
        elif brand in brand_data and "by_norm" in brand_data[brand]:
            by_norm = brand_data[brand]["by_norm"]
            n = super_norm(code)
            if n in by_norm:
                _, name, p, curr = by_norm[n]
                price = p
                currency = curr
                desc = name
                source = f"{brand} LİSTE"

        ws.append([brand, code, price, currency, desc, page, source])
        r = ws.max_row
        bc = BRAND_COLORS.get(brand, "DDDDDD").replace("#", "")
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=bc)
        ws.cell(row=r, column=1).font = Font(bold=True)
        if price in (None, "", 0):
            for col in range(2, 8):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="F2F2F2")

    widths = [14, 22, 14, 12, 50, 10, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────── UI ───────────────────────────────
if "catalog" not in st.session_state:
    st.session_state.catalog = None
if "brand_data" not in st.session_state:
    st.session_state.brand_data = {}
if "oz_settings" not in st.session_state:
    st.session_state.oz_settings = {
        "discount": 0.25,   # %75 iskonto → ×0.25
        "otv":      1.067,  # ÖTV
        "margin":   2.03,   # marj
    }

tabs = st.tabs([
    "1️⃣ Katalog",
    "2️⃣ Marka Listeleri",
    "3️⃣ Öztiryakiler Ayarları",
    "4️⃣ Önizleme",
    "5️⃣ Üret",
])

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 1. Katalog ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tabs[0]:
    st.subheader("Katalog PDF'ini yükle")
    st.caption("Bu PDF, fiyatların basılacağı kaynak kataloğdur.")
    pdf_file = st.file_uploader("PDF dosyası", type=["pdf"], key="pdf_up")

    if pdf_file is not None:
        with st.spinner("PDF taranıyor — kodlar, sayfa konumları, audit..."):
            st.session_state.catalog = scan_catalog(pdf_file.getvalue())
            st.session_state.catalog["pdf_name"] = pdf_file.name

    cat = st.session_state.catalog
    if cat is not None:
        st.success(f"✓ {cat['pdf_name']} • {cat['n_pages']} sayfa • "
                   f"{len(cat['codes_pages'])} unique ürün kodu")
        # Marka dağılımı
        brands = Counter(brand_of(c) for c in cat["codes_pages"])
        st.markdown("**Marka dağılımı:**")
        cols = st.columns(len(brands))
        for col, (b, n) in zip(cols, brands.most_common()):
            with col:
                st.metric(b, n)

        # Audit bilgisi
        c1, c2, c3 = st.columns(3)
        c1.metric("OCR Typo", len(cat["typo_pairs"]),
                  delta="problem" if cat["typo_pairs"] else "temiz",
                  delta_color="inverse")
        c2.metric("Tablo-only kod", len(cat["table_only"]),
                  delta="fiyatlanamaz" if cat["table_only"] else "temiz",
                  delta_color="inverse")
        c3.metric("Üst üste yığın", len(cat["stacks"]),
                  delta="algoritma çözer")

        with st.expander("Kod listesi (tüm sayfa+marka eşleşmeleri)"):
            rows = [{"Marka": brand_of(c), "Kod": c, "Sayfa": p}
                    for c, p in sorted(cat["codes_pages"].items(),
                                       key=lambda x: (brand_of(x[0]), x[1]))]
            st.dataframe(rows, use_container_width=True, hide_index=True)

        if cat["typo_pairs"]:
            with st.expander(f"⚠️ {len(cat['typo_pairs'])} OCR typo — PDF'te düzeltilebilir"):
                for canon, typo in cat["typo_pairs"]:
                    st.text(f"  Kanonik '{canon}'  ↔  Typo '{typo}'")

        if cat["table_only"]:
            with st.expander(f"⚠️ {len(cat['table_only'])} kod sadece tabloda — fiyat basılamaz"):
                for c in cat["table_only"]:
                    st.text(f"  {c}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 2. Marka Listeleri ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tabs[1]:
    st.subheader("Marka bazlı fiyat listeleri")
    st.caption("Her marka için ayrı Excel yükle. Yüklemediğin markalar boş bırakılır.")

    brand_specs = [
        ("İNOKSAN", "ZMD-XXX kodları — İnoksan listesi (MARKA KODU, STOK KODU formatı)"),
        ("ATALAY", "MD.7A kodları — M&D Prime listesi"),
        ("EMPERO", "ZMD.XXX kodları"),
        ("SEAMAC", "MD.XXX (diğer) kodları"),
        ("ÖZTİRYAKİLER", "MDK4/MDE4 kodları — özel mapping (NMV↔MDB, LMV↔MDD, PZC, prefix 79K4/79E4)"),
    ]

    for brand, hint in brand_specs:
        cols = st.columns([1, 4])
        with cols[0]:
            st.markdown(
                f"<div style='background:{BRAND_COLORS[brand]};padding:8px 12px;"
                f"border-radius:6px;font-weight:bold;text-align:center'>{brand}</div>",
                unsafe_allow_html=True,
            )
        with cols[1]:
            uploaded = st.file_uploader(
                hint, type=["xlsx"], key=f"brand_{brand}",
                label_visibility="collapsed",
            )
            if uploaded:
                try:
                    st.session_state.brand_data[brand] = index_brand_list(
                        uploaded.getvalue(), brand
                    )
                    if brand == "ÖZTİRYAKİLER":
                        n = sum(len(v) for v in st.session_state.brand_data[brand]["oz_by_sig"].values())
                        st.success(f"✓ {n} Öz kodu indekslendi")
                    else:
                        n = len(st.session_state.brand_data[brand]["by_norm"])
                        st.success(f"✓ {n} kod indekslendi")
                except Exception as e:
                    st.error(f"Okuma hatası: {e}")

    if st.session_state.brand_data:
        st.markdown("---")
        st.markdown("**Yüklenen markalar:** " + ", ".join(st.session_state.brand_data.keys()))

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 3. Öz Ayarlar ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tabs[2]:
    st.subheader("Öztiryakiler fiyat hesap parametreleri")
    st.caption("M&D liste fiyatı = Öz fiyatı × iskonto × ÖTV × marj")

    c1, c2, c3 = st.columns(3)
    with c1:
        disc_pct = st.slider("İskonto (%)", 0, 90, 75, step=1,
                             help="Öztiryakiler listesinden iskonto oranı")
        oz_disc = (100 - disc_pct) / 100.0
        st.caption(f"Çarpan: × {oz_disc:.3f}")
    with c2:
        oz_otv = st.number_input("ÖTV çarpanı", 1.000, 2.000, 1.067, 0.001, format="%.3f")
        st.caption("Tipik: 1.067 (ÖTV %6.7)")
    with c3:
        oz_margin = st.number_input("Marj çarpanı", 1.000, 5.000, 2.03, 0.01, format="%.2f")
        st.caption("Maliyet → Liste fiyatı")

    st.session_state.oz_settings = {
        "discount": oz_disc,
        "otv": oz_otv,
        "margin": oz_margin,
    }

    # Örnek hesap
    example = 1000.0
    final = example * oz_disc * oz_otv * oz_margin
    total_mult = oz_disc * oz_otv * oz_margin
    st.markdown(f"**Örnek hesap:**  Öz fiyatı **1.000 EUR** → M&D liste **{final:.2f} EUR**  "
                f"(toplam çarpan ×{total_mult:.4f})")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 4. Önizleme ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tabs[3]:
    st.subheader("Fiyat eşleştirme önizleme")
    cat = st.session_state.catalog
    if cat is None:
        st.info("Önce **1️⃣ Katalog** sekmesinden PDF yükle.")
    else:
        with st.spinner("Fiyatlar eşleştiriliyor..."):
            xlsx_bytes = build_price_excel(
                cat["codes_pages"],
                st.session_state.brand_data,
                **st.session_state.oz_settings,
            )
        # Excel'i geçici tut
        st.session_state.current_xlsx = xlsx_bytes
        # Yüklemiş gibi okuyup tablo göster
        p = _write_temp(xlsx_bytes, ".xlsx")
        wb = load_workbook(p, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({
                "Marka": row[0],
                "Kod": row[1],
                "Fiyat": row[2],
                "PB": row[3],
                "Açıklama": (row[4] or "")[:40],
                "Sayfa": row[5],
                "Kaynak": row[6],
            })

        filled = sum(1 for r in rows if r["Fiyat"] not in (None, "", 0))
        empty = len(rows) - filled

        c1, c2 = st.columns(2)
        c1.metric("Dolu fiyat", filled)
        c2.metric("Boş", empty, delta_color="inverse")

        # Marka bazlı doluluk
        st.markdown("**Marka bazlı doluluk:**")
        brand_full = Counter(); brand_total = Counter()
        for r in rows:
            brand_total[r["Marka"]] += 1
            if r["Fiyat"] not in (None, "", 0):
                brand_full[r["Marka"]] += 1
        cols = st.columns(len(brand_total))
        for col, b in zip(cols, sorted(brand_total.keys())):
            with col:
                f = brand_full[b]
                t = brand_total[b]
                pct = f * 100 // t if t else 0
                st.metric(b, f"{f}/{t}", delta=f"%{pct}",
                          delta_color="normal" if pct >= 90 else "inverse")

        # Tablo
        st.dataframe(rows, use_container_width=True, hide_index=True, height=450)

        st.download_button(
            "📥 Excel olarak indir",
            data=xlsx_bytes,
            file_name="fiyat-listesi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━ 5. Üret ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tabs[4]:
    st.subheader("Fiyatlı PDF üret")
    cat = st.session_state.catalog
    if cat is None:
        st.info("Önce **1️⃣ Katalog** sekmesinden PDF yükle.")
    elif not st.session_state.brand_data:
        st.warning("**2️⃣ Marka Listeleri** sekmesinden en az bir marka listesi yükle.")
    else:
        st.markdown("Hazırsan butona bas → fiyatlı PDF üretilir.")

        if st.button("🚀 Fiyatlı PDF Üret", type="primary", use_container_width=True):
            with st.spinner("Fiyatlar hesaplanıyor ve PDF üretiliyor..."):
                xlsx_bytes = build_price_excel(
                    cat["codes_pages"],
                    st.session_state.brand_data,
                    **st.session_state.oz_settings,
                )
                xlsx_path = _write_temp(xlsx_bytes, ".xlsx")
                prices = load_prices(xlsx_path)

                # Fuzzy match (OCR typo'lu kodlar için)
                pdf_codes = set(cat["codes_pages"].keys())
                fuzzy = apply_fuzzy_matches(prices, pdf_codes, max_distance=1)

                # Overlay + merge
                overlay_bytes, log, stats = build_overlay(cat["pdf_path"], prices, None)
                out_path = Path(tempfile.mktemp(suffix=".pdf"))
                merge(cat["pdf_path"], overlay_bytes, out_path)
                out_bytes = out_path.read_bytes()

            total = sum(len(v) for v in log.values())
            st.success(f"✓ {total} fiyat basıldı ({len(log)} sayfada)")
            st.markdown("**Konum dağılımı:** " +
                        " · ".join(f"{k}: {v}" for k, v in stats.items()))

            if fuzzy:
                with st.expander(f"Fuzzy eşleşme: {len(fuzzy)} OCR typo düzeltildi"):
                    for pc, ec, d in fuzzy:
                        st.text(f"  PDF '{pc}' ↔ Excel '{ec}'")

            base = Path(cat["pdf_name"]).stem
            st.download_button(
                "📥 Fiyatlı PDF'i indir",
                data=out_bytes,
                file_name=f"{base} - FIYATLI.pdf",
                mime="application/pdf",
                type="primary",
                use_container_width=True,
            )
            st.download_button(
                "📥 Fiyat-listesi.xlsx (kullanılan veri)",
                data=xlsx_bytes,
                file_name="fiyat-listesi.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━ Bilgi ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.divider()
with st.expander("ℹ️ Nasıl kullanılır"):
    st.markdown(
        """
        1. **Katalog** sekmesinden PDF'i yükle. Sistem kodları çıkarır, audit yapar.
        2. **Marka Listeleri**'nden her markaya ait Excel'i yükle.
        3. **Öztiryakiler Ayarları**'nda iskonto/ÖTV/marj parametrelerini ayarla.
        4. **Önizleme**'den fiyat eşleşmesini kontrol et, boş kalan varsa söyler.
        5. **Üret**'e bas → fiyatlı PDF'i indir.

        **Otomatik özellikler:**
        - Fiyat ürün başlığı altına basılır (üstüste yığın varsa kodun soluna)
        - Spec tablosu satırlarına dokunmaz
        - Renkli arka plana düşerse pozisyonu ayarlar
        - OCR typo'lu kodları otomatik düzeltir
        - Öztiryakiler kodları için: prefix öncelik (79K4, 79E4, ...)
          ve mapping (NMV↔MDB, LMV↔MDD)
        """
    )
