"""
Excel'deki fiyatları PDF'in üzerine overlay olarak basar.
Orijinal tasarım korunur — sadece fiyat metni eklenir.

Kullanım:
    python3 inject_prices.py                       # tüm katalog
    python3 inject_prices.py --pages 9-12          # sadece bu sayfalar (preview)
    python3 inject_prices.py --pages 9-12 --dummy  # excel boşsa dummy fiyat ile
"""
from __future__ import annotations
import argparse
import io
import re
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas

HERE = Path(__file__).parent
PDF_IN = HERE / "MD CATALOG 2025 TR-EN.pdf"
XLSX = HERE / "fiyat-listesi.xlsx"
PDF_OUT_DEFAULT = HERE / "MD CATALOG 2025 - FIYATLI.pdf"

CODE_RE = re.compile(
    r"\b(?:ZMD|MD)[.\-][A-Z0-9][A-Z0-9.\-/]*[A-Z0-9]"
    r"(?:\s+\d[A-Z0-9.\-/]*[A-Z][A-Z0-9.\-/]*)?\b"
)

# Kullanıcı PDF'e [FIYAT] / [FİYAT] placeholder'ı yerleştirebilir. Bu marker
# bulunduğunda en yakın kodun fiyatı buraya basılır (otomatik yerleştirme devre dışı).
# Türkçe İ ile İngilizce I'nın her ikisini de tolere eder.
MARKER_RE = re.compile(r"\[[Ff][IİIıi][Yy][Aa][Tt]\]")

# Overlay görünümü
PRICE_COLOR = HexColor("#C00000")  # koyu kırmızı
PRICE_FONT = "Helvetica-Bold"
PRICE_SIZE = 9
PRICE_LINE_GAP = -1  # kodun alt kenarı ile fiyat üst kenarı arasındaki boşluk (pt)


def parse_pages(spec: str | None, total: int) -> set[int] | None:
    """'9-12,15' -> {9,10,11,12,15}. None ise hepsi."""
    if not spec:
        return None
    out: set[int] = set()
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            a, b = part.split("-", 1)
            out.update(range(int(a), int(b) + 1))
        else:
            out.add(int(part))
    return {p for p in out if 1 <= p <= total}


def load_prices(xlsx_path: Path) -> dict[str, str]:
    """Kod -> 'fiyat para_birimi' (formatlı)."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    out: dict[str, str] = {}
    headers = {str(c.value).strip().lower(): c.column for c in ws[1] if c.value}
    code_col = headers.get("ürün kodu") or headers.get("urun kodu") or 1
    price_col = headers.get("fiyat") or 2
    curr_col = headers.get("para birimi") or 3
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[code_col - 1]
        price = row[price_col - 1]
        curr = row[curr_col - 1] if len(row) >= curr_col else "TL"
        if not code or price in (None, ""):
            continue
        try:
            p = float(price)
            formatted = f"{p:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except (TypeError, ValueError):
            formatted = str(price)
        out[str(code).strip()] = f"{formatted} {curr or 'TL'}"
    return out


def build_dummy_prices(pdf_path: Path) -> dict[str, str]:
    """Excel doldurulmadan önce preview için."""
    prices: dict[str, str] = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for m in CODE_RE.finditer(text):
                code = m.group(0)
                if code not in prices:
                    base = 1000 + (hash(code) % 9000)
                    prices[code] = f"{base:,}".replace(",", ".") + ",00 TL"
    return prices


def _edit_distance(a: str, b: str, threshold: int = 2) -> int:
    """Erken çıkışlı edit distance — eşiği aşan değerleri eşik+1 olarak döner."""
    la, lb = len(a), len(b)
    if abs(la - lb) > threshold:
        return threshold + 1
    if la > lb:
        a, b = b, a
        la, lb = lb, la
    prev = list(range(lb + 1))
    for i in range(1, la + 1):
        cur = [i]
        row_min = i
        for j in range(1, lb + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            v = min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + cost)
            cur.append(v)
            if v < row_min:
                row_min = v
        if row_min > threshold:
            return threshold + 1
        prev = cur
    return prev[lb]


def _is_ocr_typo_diff(a: str, b: str) -> bool:
    """İki kod arasındaki tek fark bir OCR-confusion pair ise (0↔O, 1↔I gibi)
    True döner. Bu fuzzy match için 'gerçek typo' olduğunu garantiler — ürün
    varyantı (S/T/W gibi suffix) olarak yanlış pozitif vermez."""
    OCR_PAIRS = {("0", "O"), ("1", "I"), ("1", "l"), ("2", "Z"),
                 ("5", "S"), ("6", "G"), ("7", "T"), ("8", "B")}
    if len(a) != len(b):
        return False
    diffs = [(ca, cb) for ca, cb in zip(a, b) if ca != cb]
    if len(diffs) != 1:
        return False
    ca, cb = diffs[0]
    return (ca, cb) in OCR_PAIRS or (cb, ca) in OCR_PAIRS


def apply_fuzzy_matches(prices: dict, pdf_codes: set, max_distance: int = 1) -> list[tuple[str, str, int]]:
    """Excel'de fiyatı olan ama PDF'te bulunamayan bir kod için PDF'teki yakın
    bir kodu bağlar. Sadece OCR-confusion typo'ları (0↔O, 1↔I vb.) eşleştirir;
    ürün varyantları (örn. ZMD-7KE10 vs ZMD-7KE10S) yanlış pozitif vermez.
    `prices` in-place genişletilir."""
    if not prices:
        return []
    excel_unmatched = [e for e in prices if e not in pdf_codes]
    used: set[str] = set()
    matched: list[tuple[str, str, int]] = []
    for pdf_code in pdf_codes:
        if pdf_code in prices:
            continue
        best_match = None
        for ec in excel_unmatched:
            if ec in used:
                continue
            if _is_ocr_typo_diff(pdf_code, ec):
                best_match = ec
                break
        if best_match is not None:
            prices[pdf_code] = prices[best_match]
            used.add(best_match)
            matched.append((pdf_code, best_match, 1))
    return matched


def find_code_positions(page) -> list[tuple[str, float, float, float, float]]:
    """Sayfadaki her kod için (kod, x0, y0, x1, y1) döner.
    pdfplumber Y'si yukarıdan; reportlab Y'si aşağıdan — çağıran çevirir.

    Boşluklu kodlar (örn. "ZMD.BYK 270L-R") iki ayrı kelime olarak okunur;
    bitişik aynı-satır kelimeleri tek kod halinde birleştirir.
    """
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    words_sorted = sorted(enumerate(words), key=lambda iw: (iw[1]["top"], iw[1]["x0"]))
    positions: list[tuple[str, float, float, float, float]] = []
    consumed: set[int] = set()

    # 1) Tek kelime + bitişik yan kelime birleşimini dene
    for idx, w in words_sorted:
        if idx in consumed:
            continue
        txt = w["text"]
        if not CODE_RE.fullmatch(txt):
            continue
        # Bitişik bir sonraki kelime ile birleşim CODE_RE'ye uyuyor mu?
        combined_added = False
        for jdx, w2 in words_sorted:
            if jdx == idx or jdx in consumed:
                continue
            same_y = abs((w2["top"] + w2["bottom"]) / 2 - (w["top"] + w["bottom"]) / 2) <= 4
            if not same_y:
                continue
            gap = w2["x0"] - w["x1"]
            if not (0 <= gap <= 15):
                continue
            combined = f"{txt} {w2['text']}"
            if CODE_RE.fullmatch(combined):
                positions.append((
                    combined,
                    w["x0"],
                    min(w["top"], w2["top"]),
                    w2["x1"],
                    max(w["bottom"], w2["bottom"]),
                ))
                consumed.add(idx)
                consumed.add(jdx)
                combined_added = True
                break
        if not combined_added:
            positions.append((txt, w["x0"], w["top"], w["x1"], w["bottom"]))
            consumed.add(idx)

    return positions


# Tek başına geçebilen birim kelimeleri (pdfplumber "2000 mm" → "2000" + "mm")
_UNIT_TOKENS = {
    "mm", "cm", "m", "kg", "kw", "kw/h", "lt", "lt/h", "v", "hz", "w", "inch",
    "adet", "psc", "pcs", "sn", "sec", "ad", "°c", "m³/h",
}
_UNIT_RE = re.compile(
    r"\b\d+(?:[.,]\d+)?\s*(?:mm|cm|m|kg|kw|lt|v|hz|w|°c|°|inch|adet|psc|pcs|sn|sec|ad)\b",
    re.IGNORECASE,
)
_FRACTION_RE = re.compile(r"^[123]/[234]\"?$")


_PURE_NUMBER_RE = re.compile(r"^\d+(?:[.,]\d+)?$")


def _is_table_row(pos, all_words, max_gap: float = 100.0) -> bool:
    """Kodun hemen sağında (max_gap pt içinde) EN AZ 2 sayı/birim/kesir kelimesi
    yan yana varsa, bu bir spec-tablo satırıdır. Tek bir yakın sayı, yan-sütun
    başlığının ilk kelimesi olabilir — tablo sayılmaz."""
    _, x0, y_top, x1, y_bot = pos
    cy = (y_top + y_bot) / 2
    right_words = [
        w for w in all_words
        if abs((w["top"] + w["bottom"]) / 2 - cy) <= 6
        and w["x0"] > x1
        and w["x0"] - x1 <= max_gap
    ]
    matches = 0
    for w in right_words:
        t = w["text"].strip()
        if (t.lower() in _UNIT_TOKENS or _UNIT_RE.search(t)
                or _FRACTION_RE.match(t) or _PURE_NUMBER_RE.match(t)):
            matches += 1
    return matches >= 2


def _select_title_positions(all_positions, prices, page):
    """Her kod için ürün başlığı konumunu seçer; spec tablosundaki satırları eler.
    Eğer bir kodun tüm konumları tablo satırı ise o kod hiç basılmaz."""
    all_words = list(page.extract_words())

    pos_by_code: dict[str, list[tuple]] = {}
    for pos in all_positions:
        if pos[0] in prices:
            pos_by_code.setdefault(pos[0], []).append(pos)

    chosen: dict[str, tuple] = {}
    for code, positions in pos_by_code.items():
        for pos in positions:
            if not _is_table_row(pos, all_words):
                chosen[code] = pos
                break
        # Tüm konumlar tablo satırıysa: skip (fiyat basma).

    ordered = []
    seen = set()
    for pos in all_positions:
        code = pos[0]
        if code in chosen and code not in seen:
            ordered.append(chosen[code])
            seen.add(code)
    return ordered


def _overlap_area(a, b) -> float:
    aw = max(0.0, min(a[2], b[2]) - max(a[0], b[0]))
    ah = max(0.0, min(a[3], b[3]) - max(a[1], b[1]))
    return aw * ah


def _is_white_color(color) -> bool:
    if color is None:
        return True
    if isinstance(color, (int, float)):
        return color > 0.95
    if isinstance(color, (tuple, list)):
        if len(color) == 1:
            return color[0] > 0.95
        if len(color) == 3:
            return all(c > 0.95 for c in color)
        if len(color) == 4:
            return all(c < 0.05 for c in color[:3])
    return False


def _is_dark_color(color) -> bool:
    """Kırmızı fiyat ile kontrastı zayıf olan renkler için beyaz arkaplan gerekir."""
    if color is None:
        return False
    if isinstance(color, (int, float)):
        return color < 0.7
    if isinstance(color, (tuple, list)):
        if len(color) == 1:
            return color[0] < 0.7
        if len(color) == 3:
            lum = 0.299 * color[0] + 0.587 * color[1] + 0.114 * color[2]
            return lum < 0.7
        if len(color) == 4:
            return sum(color[:3]) > 0.9
    return False


def _get_colored_fills(page):
    """Renkli (beyaz olmayan) dolgu shape'leri: (x0, top, x1, bot, color)."""
    out = []
    pw, ph = page.width, page.height
    for shape in list(page.rects) + list(page.curves):
        if not shape.get("fill"):
            continue
        color = shape.get("non_stroking_color")
        if _is_white_color(color):
            continue
        sw = shape.get("width", 0) or 0
        sh = shape.get("height", 0) or 0
        if sw < 3 or sh < 3:
            continue
        if sw * sh > pw * ph * 0.5:
            continue
        out.append((shape["x0"], shape["top"], shape["x1"], shape["bottom"], color))
    return out


def _overlapping_fill_color(bbox, colored_fills):
    """Bu bbox renkli bir dolguyla çakışıyorsa, dominant rengini döner."""
    best_color = None
    best_area = 0.0
    for f in colored_fills:
        ow = max(0.0, min(bbox[2], f[2]) - max(bbox[0], f[0]))
        oh = max(0.0, min(bbox[3], f[3]) - max(bbox[1], f[1]))
        a = ow * oh
        if a > best_area:
            best_area = a
            best_color = f[4]
    return best_color


def _get_obstacles(page, exclude_bbox=None):
    """Sayfadaki engelleri (x0, top, x1, bottom) listesi olarak döner:
    metin kelimeleri + dolgu içeren dikdörtgenler (tablo başlığı çubukları vb.) + görseller.
    `exclude_bbox` verilirse o alana büyük oranda denk gelen kelimeler atlanır (ürün kodunun kendisi)."""
    obs: list[tuple[float, float, float, float]] = []
    pw, ph = page.width, page.height

    for w in page.extract_words():
        bb = (w["x0"], w["top"], w["x1"], w["bottom"])
        if exclude_bbox is not None:
            ba = max(1.0, (bb[2] - bb[0]) * (bb[3] - bb[1]))
            if _overlap_area(bb, exclude_bbox) / ba > 0.5:
                continue
        obs.append(bb)

    # rects + yuvarlatılmış köşeli kutular (curves) + çizgiler — tablo başlığı
    # çubukları çoğunlukla rounded-rect olduğu için "curve" olarak çıkar.
    for shape in list(page.rects) + list(page.curves) + list(page.lines):
        sw = shape.get("width", 0) or 0
        sh = shape.get("height", 0) or 0
        if sw < 3 or sh < 3:
            continue
        if sw * sh > pw * ph * 0.5:
            continue
        obs.append((shape["x0"], shape["top"], shape["x1"], shape["bottom"]))

    for img in page.images:
        obs.append((img["x0"], img["top"], img["x1"], img["bottom"]))

    # Dekoratif sayfa-boyu dikey şeritler (kırmızı kenar bandı vs.) sadece
    # sol kenarı yakalanıyor; şeridin SAĞINDA kalan alan da kullanılamaz —
    # şeridin solundan sayfa kenarına kadar tüm bandı engel olarak ekle.
    for shape in list(page.rects) + list(page.curves) + list(page.lines):
        sw = shape.get("width", 0) or 0
        sh = shape.get("height", 0) or 0
        if sh < ph * 0.7 or sw > pw * 0.15:
            continue
        sx0 = shape["x0"]
        sx1 = shape["x1"]
        if sx0 > pw * 0.7:                      # sağ sidebar
            obs.append((sx0, 0, pw, ph))
        elif sx1 < pw * 0.3:                    # sol sidebar (varsa)
            obs.append((0, 0, sx1, ph))

    return obs


def _best_placement(code_bbox, text_w, text_h, page_w, page_h, obstacles):
    """En az çarpışan yerleşimi seçer. Sıralama tercih: below > above > right > left.
    Dönüş: (label, plot_x_left, plot_y_top) — pdfplumber koordinatlarında."""
    x0, y_top, x1, y_bot = code_bbox
    code_h = y_bot - y_top
    GAP_V = -1  # dikey boşluk (kodla fiyat arası — sıkı dursun)
    GAP_H = 6   # yatay boşluk
    y_mid = y_top + (code_h - text_h) / 2

    candidates = [
        ("below", x1 - text_w, y_bot + GAP_V),
        ("above", x1 - text_w, y_top - GAP_V - text_h),
        ("right", x1 + GAP_H, y_mid),
        ("left", x0 - GAP_H - text_w, y_mid),
    ]
    MARGIN = 4
    scored = []
    for idx, (label, x, yt) in enumerate(candidates):
        if x < MARGIN or x + text_w > page_w - MARGIN:
            continue
        if yt < MARGIN or yt + text_h > page_h - MARGIN:
            continue
        bbox = (x, yt, x + text_w, yt + text_h)
        score = sum(_overlap_area(bbox, o) for o in obstacles)
        scored.append((score, idx, label, x, yt))

    if not scored:
        return ("below", x1 - text_w, y_bot + GAP_V)

    # Tolerans: below'daki çakışma metin alanının %15'inden azsa yine below seç.
    # PDF render artefaktları (ince outline curves vs.) küçük sahte overlap üretebilir.
    text_area = max(1.0, text_w * text_h)
    below = next((e for e in scored if e[2] == "below"), None)
    if below is not None and below[0] <= text_area * 0.15:
        return ("below", below[3], below[4])

    scored.sort(key=lambda t: (t[0], t[1]))
    _, _, label, x, yt = scored[0]
    return (label, x, yt)


def build_overlay(
    pdf_in: Path, prices: dict[str, str], pages_filter: set[int] | None
) -> tuple[bytes, dict[int, list[str]], dict[str, int]]:
    """Overlay PDF byte'ları + sayfa başına basılan kodlar + konum istatistikleri."""
    buf = io.BytesIO()
    c: canvas.Canvas | None = None
    page_log: dict[int, list[str]] = {}
    placement_stats: dict[str, int] = {}

    with pdfplumber.open(pdf_in) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            w, h = page.width, page.height
            if c is None:
                c = canvas.Canvas(buf, pagesize=(w, h))
            else:
                c.setPageSize((w, h))

            if pages_filter is None or i in pages_filter:
                page_obstacles = _get_obstacles(page)
                page_colored_fills = _get_colored_fills(page)
                placed_boxes: list[tuple[float, float, float, float]] = []

                all_code_positions = find_code_positions(page)

                # 1) Önce manuel marker'ları ([FIYAT] placeholder'ı) bul.
                #    Greedy assignment: en yakın (marker, kod) çiftinden başla,
                #    her marker ve her kod en fazla bir kez kullanılır.
                page_words = list(page.extract_words())
                markers = [w for w in page_words if MARKER_RE.fullmatch(w["text"])]

                # Her marker en yakın koda eşlenir (Excel'de fiyatı olsun olmasın).
                # En yakın kod prices'ta yoksa o marker fiyatsız kalır ama placeholder
                # yine silinir — [FİYAT] yazısı çıktıda görünmez.
                pairs: list[tuple[float, int, str]] = []
                for mk_idx, mk in enumerate(markers):
                    mcx = (mk["x0"] + mk["x1"]) / 2
                    mcy = (mk["top"] + mk["bottom"]) / 2
                    for pos in all_code_positions:
                        code = pos[0]
                        _, x0, y_top, x1, y_bot = pos
                        cx, cy = (x0 + x1) / 2, (y_top + y_bot) / 2
                        d = ((mcx - cx) ** 2 + (mcy - cy) ** 2) ** 0.5
                        pairs.append((d, mk_idx, code))
                pairs.sort()
                used_markers: set[int] = set()
                used_codes: set[str] = set()
                matched_mk_idx: set[int] = set()
                code_to_marker: dict[str, tuple[dict, float]] = {}
                for d, mk_idx, code in pairs:
                    if mk_idx in used_markers or code in used_codes:
                        continue
                    used_markers.add(mk_idx)
                    used_codes.add(code)
                    if code in prices:
                        code_to_marker[code] = (markers[mk_idx], d)
                        matched_mk_idx.add(mk_idx)

                manually_placed_codes: set[str] = set()
                for code, (mk, _) in code_to_marker.items():
                    text = prices[code]
                    c.setFont(PRICE_FONT, PRICE_SIZE)
                    text_w = c.stringWidth(text, PRICE_FONT, PRICE_SIZE)
                    plot_x = mk["x0"]
                    plot_y_top = mk["top"]
                    baseline_y = h - plot_y_top - PRICE_SIZE

                    # Marker'ı beyaz kutuyla ört
                    pad_mk = 1
                    c.setFillColorRGB(1, 1, 1)
                    c.rect(
                        mk["x0"] - pad_mk,
                        h - mk["bottom"] - pad_mk,
                        (mk["x1"] - mk["x0"]) + 2 * pad_mk,
                        (mk["bottom"] - mk["top"]) + 2 * pad_mk,
                        stroke=0,
                        fill=1,
                    )

                    c.setFillColor(PRICE_COLOR)
                    c.drawString(plot_x, baseline_y, text)

                    placed_boxes.append((plot_x, plot_y_top, plot_x + text_w, plot_y_top + PRICE_SIZE))
                    manually_placed_codes.add(code)
                    placement_stats["marker"] = placement_stats.get("marker", 0) + 1

                # Eşleşmeyen marker'ları temizle (fiyatı olmayan koda yakın olanlar).
                # Aksi halde "[FİYAT]" PDF'te görünür kalır.
                for mk_idx, mk in enumerate(markers):
                    if mk_idx in matched_mk_idx:
                        continue
                    c.setFillColorRGB(1, 1, 1)
                    pad_mk = 1
                    c.rect(
                        mk["x0"] - pad_mk,
                        h - mk["bottom"] - pad_mk,
                        (mk["x1"] - mk["x0"]) + 2 * pad_mk,
                        (mk["bottom"] - mk["top"]) + 2 * pad_mk,
                        stroke=0,
                        fill=1,
                    )
                    placement_stats["marker_unmatched"] = placement_stats.get("marker_unmatched", 0) + 1

                # 2) Marker ile yerleştirilmemiş kodlar için otomatik algoritma
                title_positions = [
                    p for p in _select_title_positions(all_code_positions, prices, page)
                    if p[0] not in manually_placed_codes
                ]

                # Yığın algılama: alttaki kodun y_top'ı üsttekinin y_bot'una yakın
                # olmalı (iç içe geçmiş satırlarda fark negatif bile olabilir).
                STACK_GAP_MIN = -10
                STACK_GAP_MAX = 20

                def _has_code_just_below(pos):
                    _, x0, y_top, x1, y_bot = pos
                    for other in title_positions:
                        if other is pos:
                            continue
                        _, ox0, oy_top, ox1, _ = other
                        if min(x1, ox1) - max(x0, ox0) > 0 and STACK_GAP_MIN <= (oy_top - y_bot) <= STACK_GAP_MAX:
                            return True
                    return False

                for pos in title_positions:
                    code, x0, y_top, x1, y_bot = pos

                    text = prices[code]
                    c.setFont(PRICE_FONT, PRICE_SIZE)
                    text_w = c.stringWidth(text, PRICE_FONT, PRICE_SIZE)

                    # 1) Default konum: kodun ALTINA (veya yığın üstündeyse ÜSTÜNE)
                    plot_x = x1 - text_w
                    if _has_code_just_below(pos):
                        plot_y_top = y_top - PRICE_LINE_GAP - PRICE_SIZE
                        label = "above"
                    else:
                        plot_y_top = y_bot + PRICE_LINE_GAP
                        label = "below"

                    # 2) Default konum renkli bir alana mı düşüyor? Düşüyorsa
                    #    fiyatı kodun SOLUNA al.
                    bbox = (plot_x, plot_y_top, plot_x + text_w, plot_y_top + PRICE_SIZE)
                    if _overlapping_fill_color(bbox, page_colored_fills) is not None:
                        plot_x = x0 - 6 - text_w
                        plot_y_top = (y_top + y_bot - PRICE_SIZE) / 2
                        label = "left"

                    # 3) Final konum koyu bir alana düşüyor mu? Düşüyorsa beyaz
                    #    arkaplan ile kontrast sağla.
                    bbox = (plot_x, plot_y_top, plot_x + text_w, plot_y_top + PRICE_SIZE)
                    bg_color = _overlapping_fill_color(bbox, page_colored_fills)
                    needs_white_bg = bg_color is not None and _is_dark_color(bg_color)

                    placement_stats[label] = placement_stats.get(label, 0) + 1

                    baseline_y = h - plot_y_top - PRICE_SIZE

                    if needs_white_bg:
                        pad = 2
                        c.setFillColorRGB(1, 1, 1)
                        c.rect(
                            plot_x - pad,
                            baseline_y - pad,
                            text_w + 2 * pad,
                            PRICE_SIZE + 2 * pad,
                            stroke=0,
                            fill=1,
                        )

                    c.setFillColor(PRICE_COLOR)
                    c.drawString(plot_x, baseline_y, text)

                    placed_boxes.append(
                        (plot_x, plot_y_top,
                         plot_x + text_w, plot_y_top + PRICE_SIZE)
                    )

                rendered_codes = [p[0] for p in title_positions] + list(manually_placed_codes)
                if rendered_codes:
                    page_log[i] = sorted(set(rendered_codes))

            c.showPage()

    if c is not None:
        c.save()
    return buf.getvalue(), page_log, placement_stats


def merge(pdf_in: Path, overlay_bytes: bytes, pdf_out: Path) -> None:
    base = PdfReader(str(pdf_in))
    over = PdfReader(io.BytesIO(overlay_bytes))
    writer = PdfWriter()
    for i, page in enumerate(base.pages):
        if i < len(over.pages):
            page.merge_page(over.pages[i])
        writer.add_page(page)
    with open(pdf_out, "wb") as f:
        writer.write(f)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pages", help="örn. 9-12 veya 9-12,20,25")
    ap.add_argument("--dummy", action="store_true", help="Excel yerine sahte fiyat")
    ap.add_argument("--out", type=Path, default=None)
    ap.add_argument("--pdf-in", type=Path, default=None, help="Default yerine başka bir PDF kullan")
    args = ap.parse_args()

    pdf_in: Path = args.pdf_in if args.pdf_in else PDF_IN
    if not pdf_in.exists():
        raise SystemExit(f"PDF yok: {pdf_in}")

    if args.dummy:
        prices = build_dummy_prices(pdf_in)
        print(f"  (dummy fiyat: {len(prices)} kod)")
    else:
        if not XLSX.exists():
            raise SystemExit(f"Excel yok: {XLSX}. Önce extract_codes.py çalıştır.")
        prices = load_prices(XLSX)
        print(f"  (excel'den fiyatlı kod: {len(prices)})")
        if not prices:
            raise SystemExit("Excel'de fiyatlı satır yok. --dummy ile preview çıkarabilirsin.")

    with pdfplumber.open(pdf_in) as pdf:
        total = len(pdf.pages)
        # PDF'teki tüm kodları topla — Excel'de tam eşleşmesi olmayanlar için
        # fuzzy match denenecek (typo'lu kodları yakalamak için).
        pdf_codes: set[str] = set()
        for page in pdf.pages:
            text = page.extract_text() or ""
            for m in CODE_RE.finditer(text):
                pdf_codes.add(m.group(0))

    fuzzy = apply_fuzzy_matches(prices, pdf_codes, max_distance=1)
    if fuzzy:
        print(f"  OCR-typo fuzzy eşleşme ({len(fuzzy)}):")
        for pdf_code, excel_code, d in fuzzy:
            print(f"    PDF '{pdf_code}' ↔ Excel '{excel_code}'")

    # Excel'de var ama PDF'te eşleşmesi olmayan kodlar — kullanıcıyı uyar
    unmatched_excel = sorted(c for c in prices if c not in pdf_codes)
    if unmatched_excel:
        print(f"  ⚠️  Excel'de var, PDF'te bulunamadı ({len(unmatched_excel)}):")
        for c in unmatched_excel[:20]:
            print(f"    {c}")
        if len(unmatched_excel) > 20:
            print(f"    … ve {len(unmatched_excel) - 20} kod daha")

    pages_filter = parse_pages(args.pages, total)
    if pages_filter:
        print(f"  sayfalar: {sorted(pages_filter)}")

    overlay_bytes, log, stats = build_overlay(pdf_in, prices, pages_filter)
    out_path = args.out or (
        HERE / ("PREVIEW - " + pdf_in.stem + ".pdf") if pages_filter else
        HERE / (pdf_in.stem + " - FIYATLI.pdf")
    )
    merge(pdf_in, overlay_bytes, out_path)

    total_placed = sum(len(v) for v in log.values())
    print(f"✓ {total_placed} fiyat etiketi basıldı ({len(log)} sayfada)")
    if stats:
        breakdown = ", ".join(f"{k}: {v}" for k, v in sorted(stats.items(), key=lambda x: -x[1]))
        print(f"  konum dağılımı → {breakdown}")
    print(f"✓ Çıktı: {out_path.name}")


if __name__ == "__main__":
    main()
