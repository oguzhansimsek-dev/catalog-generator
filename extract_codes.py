"""
PDF'ten ürün kodlarını çıkarıp Excel şablonu oluşturur.
Şablonu doldurduktan sonra inject_prices.py ile fiyatlı PDF üretilir.
"""
import re
from pathlib import Path
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PDF_PATH = Path(__file__).parent / "MD CATALOG 2025 TR-EN.pdf"
OUT_XLSX = Path(__file__).parent / "fiyat-listesi.xlsx"

# ZMD-7KE23, ZMD.150.70.01-2C, MD.7AKE.8085 gibi kodları yakalar.
# Opsiyonel boşluklu varyant suffix de dahil edilir (örn. "ZMD.BYK 270L-R").
CODE_RE = re.compile(
    r"\b(?:ZMD|MD)[.\-][A-Z0-9][A-Z0-9.\-/]*[A-Z0-9]"
    r"(?:\s+\d[A-Z0-9.\-/]*[A-Z][A-Z0-9.\-/]*)?\b"
)


# OCR/font kaynaklı sık karışan harf-rakam çiftleri (digit, harf).
# Tek karakterlik fark bir OCR typo ise rakamlı versiyon kanonik kabul edilir.
_OCR_PAIRS: dict[str, str] = {
    "0": "O", "1": "I", "1": "l", "2": "Z", "5": "S", "6": "G", "7": "T", "8": "B",
}


def _pick_typo_winner(a: str, b: str) -> tuple[str, str] | None:
    """İki kod tek bir OCR-karışıklığı (rakam ↔ benzer harf) ile farklıysa
    (kanonik, typo) döner. Aksi halde None — gerçek farklı kod olarak kalır."""
    if len(a) != len(b):
        return None
    diffs = [(i, ca, cb) for i, (ca, cb) in enumerate(zip(a, b)) if ca != cb]
    if len(diffs) != 1:
        return None
    _, ca, cb = diffs[0]
    if ca in _OCR_PAIRS and cb == _OCR_PAIRS[ca]:
        return (a, b)
    if cb in _OCR_PAIRS and ca == _OCR_PAIRS[cb]:
        return (b, a)
    return None


def _dedupe_typos(codes_pages: dict[str, int]) -> dict[str, int]:
    """Tipo görünen kodları kanonik kodla birleştirir."""
    code_list = list(codes_pages.keys())
    drop: set[str] = set()
    for i, a in enumerate(code_list):
        if a in drop:
            continue
        for b in code_list[i + 1:]:
            if b in drop:
                continue
            winner = _pick_typo_winner(a, b)
            if winner is None:
                continue
            _, typo = winner
            drop.add(typo)
    if drop:
        print(f"  (OCR typo eler: {len(drop)} → {sorted(drop)})")
    return {c: p for c, p in codes_pages.items() if c not in drop}


def extract_codes_with_pages(pdf_path: Path):
    """Her kodun ilk göründüğü sayfayı tutar, sırayı korur.
    OCR kaynaklı typo'lu varyantlar elenir (kanonik rakam-içeren versiyon tutulur)."""
    seen: dict[str, int] = {}
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for m in CODE_RE.finditer(text):
                code = m.group(0)
                if code not in seen:
                    seen[code] = i
    return _dedupe_typos(seen)


def build_template(codes: dict[str, int], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fiyat Listesi"

    headers = ["Ürün Kodu", "Fiyat", "Para Birimi", "Açıklama (opsiyonel)", "Sayfa (PDF)"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for code, page in codes.items():
        ws.append([code, None, "TL", "", page])

    widths = [22, 14, 14, 40, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(out_path)


def main():
    if not PDF_PATH.exists():
        raise SystemExit(f"PDF bulunamadı: {PDF_PATH}")
    codes = extract_codes_with_pages(PDF_PATH)
    build_template(codes, OUT_XLSX)
    print(f"✓ {len(codes)} adet ürün kodu bulundu.")
    print(f"✓ Şablon yazıldı: {OUT_XLSX.name}")
    print("\nİlk 10 kod:")
    for code, page in list(codes.items())[:10]:
        print(f"  {code:<24} (sayfa {page})")


if __name__ == "__main__":
    main()
